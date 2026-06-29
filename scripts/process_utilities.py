"""
水電瓦斯數據整理腳本

用法：
    python3 process.py <pdf路徑> <範例資料夾路徑> <輸出資料夾路徑> [民國年]

例：
    python3 process.py 恩博格114水電瓦斯.pdf ../範例 . 114
"""

import sys
import os
import re
import copy
import shutil
import calendar
import zipfile
import pdfplumber
from openpyxl import load_workbook

# ── 設定 ──────────────────────────────────────────────────────────────────────
TEMPLATE_MAP = {
    '電費': '電-台電.xlsx',
    '水費': '水-用水.xlsx',
    '瓦斯': '氣-天然氣、瓦斯.xlsx',
}

def get_billing_period(billing_month, year=2025):
    """依奇偶月規則計算計費日期區間。
    奇數月：該月1日 ~ 下月最後一日
    偶數月：前月1日 ~ 該月最後一日
    """
    if billing_month % 2 == 1:
        start_m = billing_month
        end_m   = billing_month + 1
    else:
        start_m = billing_month - 1
        end_m   = billing_month

    last_day = calendar.monthrange(year, end_m)[1]
    start = f"{year}-{start_m:02d}-01"
    end   = f"{year}-{end_m:02d}-{last_day:02d}"
    return f'["{start}","{end}"]'

def categorize(summary):
    """將摘要文字分類為 電費/水費/瓦斯"""
    if '電費' in summary:
        return '電費'
    if '水費' in summary:
        return '水費'
    if '天然氣' in summary or '瓦斯' in summary:
        return '瓦斯'
    return None

def extract_billing_month(summary):
    """從 '03月份電費' 之類的字串中取出月份數字"""
    m = re.match(r'(\d{1,2})月', summary)
    return int(m.group(1)) if m else None

SKIP_SUMMARIES = {'摘 要', '承上頁', '過次頁', '', '本月合計', '本期累計'}

def detect_columns(table):
    """
    從表格前幾列自動偵測欄位位置。
    回傳 (col_month, col_day, col_summary, col_debit)。
    預設值：(2, 3, 5, 7)
    """
    col_m, col_d, col_s, col_b = 2, 3, 5, 7
    for row in table[:6]:
        cells = [str(c or '').strip() for c in row]
        for i, c in enumerate(cells):
            if c in ('月',):     col_m = i
            if c in ('日',):     col_d = i
            if '摘' in c and '要' in c: col_s = i
            if '借' in c and '方' in c and '額' in c: col_b = i
    return col_m, col_d, col_s, col_b

def extract_data(pdf_path):
    """從 PDF 擷取所有資料列，回傳 [(月, 日, 摘要, 借方金額), ...]"""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                col_m, col_d, col_s, col_b = detect_columns(table)
                prev_summary = None
                for row in table:
                    n = len(row)
                    # 摘要列
                    s_val = str(row[col_s] or '').strip() if col_s < n else ''
                    if s_val and s_val not in SKIP_SUMMARIES:
                        prev_summary = s_val
                    # 資料列（有月/日/借方）
                    m_val = str(row[col_m] or '').strip() if col_m < n else ''
                    d_val = str(row[col_d] or '').strip() if col_d < n else ''
                    b_val = str(row[col_b] or '').strip().replace(',', '') if col_b < n else ''
                    if (m_val.isdigit() and d_val.isdigit() and
                            1 <= int(m_val) <= 12 and 1 <= int(d_val) <= 31 and
                            b_val.isdigit() and prev_summary):
                        rows.append((int(m_val), int(d_val), prev_summary, int(b_val)))
    return rows

def copy_style(src, dst):
    """複製單一儲存格的格式（直接複製 _style index）"""
    dst._style = copy.copy(src._style)

def fix_xlsx_for_upload(xlsx_path, template_path, ref_row=8):
    """
    openpyxl 存檔後的 XML 後處理，解決上傳系統三個驗證問題：

    1. inlineStr → shared strings
       openpyxl 預設寫 t="inlineStr"，上傳系統要求 t="s"（shared strings）。
       修補：將所有 <is><t>value</t></is> 轉成 sharedStrings 索引 <v>N</v>。

    2. 空白格 t="n" → 自閉合 <c .../>
       openpyxl 對空值格寫出 <c ... t="n"></c>，與範例的自閉合格式不符。

    3. 列層級 s 屬性
       openpyxl 存檔時覆寫 <row> 的 s/customFormat/ht/customHeight，
       從範例 row 8 取原始屬性補回。
    """
    # ── 讀取所有檔案 ────────────────────────────────────────────────────────
    with zipfile.ZipFile(xlsx_path) as z:
        all_files = {n: z.read(n) for n in z.namelist()}

    sheet_name = next(n for n in all_files if re.match(r'xl/worksheets/sheet\d+\.xml', n))
    sheet_xml  = all_files[sheet_name].decode('utf-8')

    # ── 讀取 sharedStrings ──────────────────────────────────────────────────
    SS_NAME = 'xl/sharedStrings.xml'
    if SS_NAME in all_files:
        ss_xml  = all_files[SS_NAME].decode('utf-8')
        sis     = re.findall(r'<si>(.*?)</si>', ss_xml, re.DOTALL)
        strings = [''.join(re.findall(r'<t[^>]*>(.*?)</t>', si, re.DOTALL)) for si in sis]
    else:
        ss_xml  = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
                   ' count="0" uniqueCount="0"></sst>')
        strings = []
    str_idx = {s: i for i, s in enumerate(strings)}

    def get_or_add(s):
        if s not in str_idx:
            str_idx[s] = len(strings)
            strings.append(s)
        return str_idx[s]

    # ── 修補 1：inlineStr → shared string ───────────────────────────────────
    def conv_inline(m):
        attrs = m.group(1)
        value = m.group(2)
        idx   = get_or_add(value)
        attrs = re.sub(r'\s*t="inlineStr"', '', attrs)
        return f'<c{attrs} t="s"><v>{idx}</v></c>'

    sheet_xml = re.sub(
        r'<c([^>]*)\s+t="inlineStr"><is><t[^>]*>(.*?)</t></is></c>',
        conv_inline, sheet_xml, flags=re.DOTALL
    )

    # ── 修補 2：空白格 t="n" → 自閉合 ──────────────────────────────────────
    sheet_xml = re.sub(r'<c([^>]*)\s+t="n"></c>', r'<c\1/>', sheet_xml)
    # 有值的數值格：移除多餘的 t="n"
    sheet_xml = re.sub(r'(<c[^>]*)\s+t="n"(><v>[^<]+</v></c>)', r'\1\2', sheet_xml)

    # ── 修補 3：列層級 s/ht/customFormat 屬性 ──────────────────────────────
    with zipfile.ZipFile(template_path) as z:
        tsname = next(n for n in z.namelist() if re.match(r'xl/worksheets/sheet\d+\.xml', n))
        txml   = z.read(tsname).decode('utf-8')

    ref_match = re.search(rf'<row r="{ref_row}"([^>]*)>', txml)
    if ref_match:
        ref_attrs = ref_match.group(1)
        ref_parts = [p.group(0) for p in [
            re.search(r'\bs="[^"]*"', ref_attrs),
            re.search(r'customFormat="[^"]*"', ref_attrs),
            re.search(r'\bht="[^"]*"', ref_attrs),
            re.search(r'customHeight="[^"]*"', ref_attrs),
        ] if p]

        def fix_row(m):
            r = int(m.group(1))
            if r < ref_row:
                return m.group(0)
            spans = re.search(r'spans="[^"]*"', m.group(2))
            parts = ([spans.group(0)] if spans else []) + ref_parts
            return f'<row r="{r}" {" ".join(parts)}>'

        sheet_xml = re.sub(r'<row r="(\d+)"([^>]*)>', fix_row, sheet_xml)

    # ── 更新 sharedStrings.xml ──────────────────────────────────────────────
    n = len(strings)
    new_sis = ''.join(f'<si><t>{s}</t></si>' for s in strings)
    ss_xml = re.sub(r'count="\d+"', f'count="{n}"', ss_xml)
    ss_xml = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{n}"', ss_xml)
    ss_xml = re.sub(r'(<sst[^>]*>).*?(</sst>)', rf'\g<1>{new_sis}\g<2>', ss_xml, flags=re.DOTALL)

    all_files[sheet_name] = sheet_xml.encode('utf-8')
    all_files[SS_NAME]    = ss_xml.encode('utf-8')

    CT_NAME = '[Content_Types].xml'
    if CT_NAME in all_files:
        ct_xml = all_files[CT_NAME].decode('utf-8')
        if 'sharedStrings' not in ct_xml:
            ct_xml = ct_xml.replace(
                '</Types>',
                '<Override PartName="/xl/sharedStrings.xml"'
                ' ContentType="application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sharedStrings+xml"/></Types>'
            )
            all_files[CT_NAME] = ct_xml.encode('utf-8')

    RELS_NAME = 'xl/_rels/workbook.xml.rels'
    if RELS_NAME in all_files:
        rels_xml = all_files[RELS_NAME].decode('utf-8')
        if 'sharedStrings' not in rels_xml:
            ids = list(map(int, re.findall(r'Id="rId(\d+)"', rels_xml)))
            next_id = max(ids) + 1 if ids else 5
            rels_xml = rels_xml.replace(
                '</Relationships>',
                f'<Relationship Id="rId{next_id}"'
                ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"'
                ' Target="sharedStrings.xml"/></Relationships>'
            )
            all_files[RELS_NAME] = rels_xml.encode('utf-8')

    with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as z:
        for n, d in all_files.items():
            z.writestr(n, d)

def write_xlsx(template_path, records, output_path, year=2025):
    """
    將資料寫入 XLSX。
    records: [(billing_month, 借方金額), ...]
    """
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb['數據填寫']

    REF      = 8  # 範例列
    ALL_COLS = list(range(1, 9))  # A~H

    ref_fuel = ws.cell(REF, 2).value  # 燃料代碼
    ref_unit = ws.cell(REF, 3).value  # 活動數據單位代碼

    # 清除所有範例列（避免舊備註殘留）
    for r in range(REF, ws.max_row + 1):
        for c in range(1, 10):
            ws.cell(r, c).value = None

    for i, (billing_month, amount) in enumerate(records):
        row          = REF + i
        billing_date = get_billing_period(billing_month, year)

        ws.cell(row, 1).value = '000'
        ws.cell(row, 2).value = ref_fuel
        ws.cell(row, 3).value = ref_unit
        ws.cell(row, 4).value = billing_date
        ws.cell(row, 5).value = None   # 單據號碼（留空）
        ws.cell(row, 6).value = None   # 買方統編（留空）
        ws.cell(row, 7).value = None   # 使用量（留空）
        ws.cell(row, 8).value = amount
        ws.cell(row, 9).value = None   # 備註（清除）

    # 複製第八列格式到所有資料列（含空白欄 E/F/G）
    for i in range(1, len(records)):
        row = REF + i
        for col in ALL_COLS:
            copy_style(ws.cell(REF, col), ws.cell(row, col))

    wb.save(output_path)

    # XML 後處理：修補 cell type、空白格、列層級格式
    fix_xlsx_for_upload(output_path, template_path, ref_row=REF)

    print(f"  → 已儲存 {output_path}（{len(records)} 列）")

def run(pdf_path, template_dir, output_dir, roc_year=114, ad_year=2025):
    """主流程"""
    folder_name = os.path.basename(os.path.dirname(os.path.abspath(pdf_path)))
    print(f"PDF: {pdf_path}")
    print(f"資料夾名稱: {folder_name}，民國年: {roc_year}")

    all_rows = extract_data(pdf_path)
    print(f"共擷取 {len(all_rows)} 筆資料")

    groups = {'電費': [], '水費': [], '瓦斯': []}
    for month, day, summary, amount in all_rows:
        cat           = categorize(summary)
        billing_month = extract_billing_month(summary)
        if cat and billing_month:
            groups[cat].append((billing_month, amount))

    for cat, records in groups.items():
        if not records:
            continue
        template_file = TEMPLATE_MAP[cat]
        template_path = os.path.join(template_dir, template_file)
        if not os.path.exists(template_path):
            print(f"找不到範本: {template_path}，跳過 {cat}")
            continue

        prefix   = os.path.splitext(template_file)[0]
        out_name = f"{prefix}_{roc_year}{folder_name}.xlsx"
        out_path = os.path.join(output_dir, out_name)

        print(f"\n處理 {cat}（{len(records)} 筆）→ {out_name}")
        write_xlsx(template_path, records, out_path, ad_year)

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    pdf_path     = sys.argv[1]
    template_dir = sys.argv[2]
    output_dir   = sys.argv[3]
    roc_year     = int(sys.argv[4]) if len(sys.argv) > 4 else 114
    ad_year      = roc_year + 1911

    os.makedirs(output_dir, exist_ok=True)
    run(pdf_path, template_dir, output_dir, roc_year, ad_year)
