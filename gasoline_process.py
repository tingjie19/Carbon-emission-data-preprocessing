"""
汽油數據整理腳本

用法：
    python3 gasoline_process.py <pdf路徑> <範例資料夾路徑> <輸出資料夾路徑> [民國年]
"""

import sys, os, re, copy, shutil, calendar, zipfile
import pdfplumber
from openpyxl import load_workbook

TEMPLATE_FILE = '油-汽油.xlsx'

def extract_data(pdf_path):
    """擷取汽油交易：[(月, 日, 借方金額), ...]"""
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                prev_summary = None
                for row in table:
                    if row[5] and row[5].strip() and row[5].strip() not in ('摘 要','承上頁','過次頁',''):
                        s = row[5].strip()
                        if s not in ('本月合計','本期累計'):
                            prev_summary = s
                    if (row[2] and str(row[2]).strip().isdigit() and
                            row[3] and str(row[3]).strip().isdigit()):
                        month = int(row[2])
                        day   = int(row[3])
                        raw   = row[7].strip().replace(',','') if row[7] and row[7].strip() else ''
                        if raw.isdigit() and prev_summary and any(kw in prev_summary for kw in ('汽油', '柴油')):
                            rows.append((month, day, int(raw)))
    return rows

def copy_style(src, dst):
    dst._style = copy.copy(src._style)

def fix_xlsx_for_upload(xlsx_path, template_path, ref_row=8):
    """
    XML 後處理，解決上傳系統驗證問題：
    1. inlineStr → shared strings (t="s")
    2. 空白格 t="n" → 自閉合
    3. 列層級 s/ht 屬性修補
    """
    with zipfile.ZipFile(xlsx_path) as z:
        all_files = {n: z.read(n) for n in z.namelist()}
    sheet_name = next(n for n in all_files if re.match(r'xl/worksheets/sheet\d+\.xml', n))
    sheet_xml  = all_files[sheet_name].decode('utf-8')
    SS_NAME = 'xl/sharedStrings.xml'
    if SS_NAME in all_files:
        ss_xml  = all_files[SS_NAME].decode('utf-8')
        sis     = re.findall(r'<si>(.*?)</si>', ss_xml, re.DOTALL)
        strings = [''.join(re.findall(r'<t[^>]*>(.*?)</t>', si, re.DOTALL)) for si in sis]
    else:
        ss_xml  = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                   '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
                   ' count="0" uniqueCount="0"></sst>')
        strings = []
    str_idx = {s: i for i, s in enumerate(strings)}

    def get_or_add(s):
        if s not in str_idx:
            str_idx[s] = len(strings); strings.append(s)
        return str_idx[s]

    def conv_inline(m):
        attrs = m.group(1); value = m.group(2)
        idx   = get_or_add(value)
        attrs = re.sub(r'\s*t="inlineStr"', '', attrs)
        return f'<c{attrs} t="s"><v>{idx}</v></c>'

    sheet_xml = re.sub(r'<c([^>]*)\s+t="inlineStr"><is><t[^>]*>(.*?)</t></is></c>', conv_inline, sheet_xml, flags=re.DOTALL)
    sheet_xml = re.sub(r'<c([^>]*)\s+t="n"></c>', r'<c\1/>', sheet_xml)
    sheet_xml = re.sub(r'(<c[^>]*)\s+t="n"(><v>[^<]+</v></c>)', r'\1\2', sheet_xml)

    with zipfile.ZipFile(template_path) as z:
        tsname = next(n for n in z.namelist() if re.match(r'xl/worksheets/sheet\d+\.xml', n))
        txml   = z.read(tsname).decode('utf-8')
    ref_match = re.search(rf'<row r="{ref_row}"([^>]*)>', txml)
    if ref_match:
        ref_attrs = ref_match.group(1)
        ref_parts = [p.group(0) for p in [
            re.search(r'\bs="[^"]*"', ref_attrs),
            re.search(r'customFormat="[^"]*"', ref_attrs),
            re.search(r'\bht="[^"]*"', ref_attrs),
            re.search(r'customHeight="[^"]*"', ref_attrs),
        ] if p]
        def fix_row(m):
            r = int(m.group(1))
            if r < ref_row: return m.group(0)
            spans = re.search(r'spans="[^"]*"', m.group(2))
            parts = ([spans.group(0)] if spans else []) + ref_parts
            return f'<row r="{r}" {" ".join(parts)}>'
        sheet_xml = re.sub(r'<row r="(\d+)"([^>]*)>', fix_row, sheet_xml)

    n = len(strings)
    new_sis = ''.join(f'<si><t>{s}</t></si>' for s in strings)
    ss_xml = re.sub(r'count="\d+"', f'count="{n}"', ss_xml)
    ss_xml = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{n}"', ss_xml)
    ss_xml = re.sub(r'(<sst[^>]*>).*?(</sst>)', rf'\g<1>{new_sis}\g<2>', ss_xml, flags=re.DOTALL)
    all_files[sheet_name] = sheet_xml.encode('utf-8')
    all_files[SS_NAME]    = ss_xml.encode('utf-8')

    CT_NAME = '[Content_Types].xml'
    if CT_NAME in all_files:
        ct_xml = all_files[CT_NAME].decode('utf-8')
        if 'sharedStrings' not in ct_xml:
            ct_xml = ct_xml.replace(
                '</Types>',
                '<Override PartName="/xl/sharedStrings.xml"'
                ' ContentType="application/vnd.openxmlformats-officedocument'
                '.spreadsheetml.sharedStrings+xml"/></Types>'
            )
            all_files[CT_NAME] = ct_xml.encode('utf-8')

    RELS_NAME = 'xl/_rels/workbook.xml.rels'
    if RELS_NAME in all_files:
        rels_xml = all_files[RELS_NAME].decode('utf-8')
        if 'sharedStrings' not in rels_xml:
            ids = list(map(int, re.findall(r'Id="rId(\d+)"', rels_xml)))
            next_id = max(ids) + 1 if ids else 5
            rels_xml = rels_xml.replace(
                '</Relationships>',
                f'<Relationship Id="rId{next_id}"'
                ' Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings"'
                ' Target="sharedStrings.xml"/></Relationships>'
            )
            all_files[RELS_NAME] = rels_xml.encode('utf-8')

    with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as z:
        for n, d in all_files.items():
            z.writestr(n, d)

def write_xlsx(template_path, records, output_path, year=2025):
    shutil.copy2(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb['數據填寫']
    REF = 8
    ALL_COLS = list(range(1, 10))  # A~I

    ref_fuel = ws.cell(REF, 2).value
    ref_unit = ws.cell(REF, 3).value
    ref_name = ws.cell(REF, 7).value  # 品名 (G)

    for r in range(REF, ws.max_row + 1):
        for c in range(1, 11):
            ws.cell(r, c).value = None

    for i, (month, day, amount) in enumerate(records):
        row = REF + i
        ws.cell(row, 1).value = '000'
        ws.cell(row, 2).value = ref_fuel
        ws.cell(row, 3).value = ref_unit
        ws.cell(row, 4).value = f'{year}-{month:02d}-{day:02d}'
        ws.cell(row, 5).value = None   # 單據號碼
        ws.cell(row, 6).value = None   # 買方統編
        ws.cell(row, 7).value = ref_name
        ws.cell(row, 8).value = None   # 公升數
        ws.cell(row, 9).value = amount

    for i in range(1, len(records)):
        row = REF + i
        for col in ALL_COLS:
            copy_style(ws.cell(REF, col), ws.cell(row, col))

    wb.save(output_path)
    fix_xlsx_for_upload(output_path, template_path, ref_row=REF)
    print(f"  → 已儲存 {output_path}（{len(records)} 列）")

def run(pdf_path, template_dir, output_dir, roc_year=114, ad_year=2025):
    folder_name = os.path.basename(os.path.dirname(os.path.abspath(pdf_path)))
    print(f"PDF: {pdf_path}")
    print(f"資料夾名稱: {folder_name}，民國年: {roc_year}")

    records = extract_data(pdf_path)
    print(f"共擷取 {len(records)} 筆汽油資料")

    template_path = os.path.join(template_dir, TEMPLATE_FILE)
    out_name  = f"油-汽油_{roc_year}{folder_name}.xlsx"
    out_path  = os.path.join(output_dir, out_name)
    write_xlsx(template_path, records, out_path, ad_year)

if __name__ == '__main__':
    if len(sys.argv) < 4:
        print(__doc__); sys.exit(1)
    pdf_path     = sys.argv[1]
    template_dir = sys.argv[2]
    output_dir   = sys.argv[3]
    roc_year     = int(sys.argv[4]) if len(sys.argv) > 4 else 114
    ad_year      = roc_year + 1911
    os.makedirs(output_dir, exist_ok=True)
    run(pdf_path, template_dir, output_dir, roc_year, ad_year)
